# encode:utf-8
from openpyxl import Workbook
import os
import numpy as np
### 1. reading results from text file
path = r'C:\Users\Administrator\Desktop\temp\data\done\\'

def file_name(file_dir):
    for root, dirs, filess in os.walk(file_dir):
        # print(root)  # 当前目录路径
        # print(dirs)  # 当前路径下所有子目录
        # print(filess)  # 当前路径下所有非目录子文件
        return filess

files = file_name(path)

f1_dict = {}
first_dict = {}
last_dict = {}

for file in files:
    f = open(path + file, encoding='UTF-8')
    flag = file.split('.')[-1]
    if flag == 'txt':
        f.close()
        continue

    elif flag == 'txt_f1':
        f1 = np.mean([float(i) for i in f.readlines()[-1].split()])
        f1_dict[file.split('.')[0]] = f1
    elif flag == 'txt_forward_performance':
        first = np.mean([float(i) for i in f.readlines()])
        first_dict[file.split('.')[0]] = first
    elif flag == 'txt_performance':
        last = np.mean([float(i) for i in f.readlines()])
        last_dict[file.split('.')[0]] = last

    f.closed


wb = Workbook()
ws = wb.active

i = 1
ws.cell(i, 1).value = 'dataset'
ws.cell(i, 2).value = 'mathod name'
ws.cell(i, 3).value = 'first acc'
ws.cell(i, 4).value = 'last acc'
ws.cell(i, 6).value = 'bwt'
ws.cell(i, 5).value = 'f1'

i+=1
for key in f1_dict:
    ws.cell(i, 1).value = key.split('_')[1]
    ws.cell(i, 2).value = key
    ws.cell(i, 3).value = first_dict[key] if key in first_dict else 0
    ws.cell(i, 4).value = last_dict[key] if key in last_dict else 0
    ws.cell(i, 6).value = (ws.cell(i, 4).value - ws.cell(i, 3).value) * 21 * 5  # when tasks = 21 (
    # last-firsh)*task_number/(task_number-1) *100%
    ws.cell(i, 5).value = f1_dict[key]
    i+=1

wb.save(r'C:\Users\Administrator\Desktop\temp\data\bert_only.xlsx')